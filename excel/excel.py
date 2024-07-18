import logging

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from models.database_model import ClockJS
from collections import defaultdict
from database.orm import SyncORM
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import requests
from config import Settings
from service.to_jira import Jira

pd.set_option('display.max_rows', 2000)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('expand_frame_repr', False)
pd.options.mode.chained_assignment = None
settings: Settings
jira_server: str


class Excel:
    days_types = {}
    correction_list = []

    @classmethod
    def create_excel(cls, months_years: dict, server: str, url: str):

        global jira_server
        jira_server = url
        global settings
        settings = Settings(_env_file=f'{server}.env')

        years = list(set(months_years.values()))
        for year in years:
            months = [i for i, j in months_years.items() if j == year]
            cls.insert_months(name=f'{settings.DOC_PATH}{settings.DOC_NAME} {str(year)}{settings.DOC_TYPE}', year=year,
                              months=months, server=server, url=url)

    @classmethod
    def insert_months(cls, name: str, year: int, months: list, server: str, url: str):

        data = pd.DataFrame(SyncORM.select_year(ClockJS, year, server, url))
        data = pd.DataFrame(data)

        # формируем список, содержащий дубликаты базового контракта
        drop_list = [ind for ind, row in data.iterrows() if row['kontrakt_name'][0] == 'О' and
                     row['event'] in ['permit', 'trip'] and
                     row['kontrakt_timetracking'] == 0]

        # удаляем
        data = data.drop(drop_list)

        # Если книги с переданным названием не существует, то тогда создаем новую
        try:
            wb = load_workbook(name)
            sheets = wb.sheetnames
            for month in months:
                if month in sheets:
                    wb.remove(wb[month])
        except FileNotFoundError:
            wb = Workbook()
            months = data.drop_duplicates(['period_month', 'period_year'])['period_month'].to_list()
            wb.remove(wb.active)
        for month in months:
            # подготавливаем df для вставки в excel
            df_all = cls.df_excel(data, month, year)

            # делаем шапку документа
            cls.doc_header(wb, month, year)

            # выгружаем df и раскрашиваем все, что нужно
            cls.beauty(wb, df_all, month, year)

        wb.save(name)
        wb.close()
        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=FINISHED CREATE EXCEL {str(year)}')

        Jira.attach(url, server, f'{settings.DOC_NAME} {str(year)}{settings.DOC_TYPE}')

        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=ATTACH EXCEL_{str(year)} '
            f'TO {jira_server.split("//")[1].upper()}')

    @classmethod
    def df_excel(cls, data, month, year):
        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=Started form a dataframe {month} {year}')

        alphabet = [chr(i) for i in range(73, 91)] + [f'A{chr(j)}' for j in range(65, 78)]
        numbers = [i for i in range(1, len(alphabet) + 1)]
        columns = dict(zip(numbers, alphabet))
        columns_dop = ([chr(i) for i in range(66, 73)] +
                       [str(i) for i in range(1, len(alphabet) + 1)] +
                       [f'A{chr(i)}' for i in range(78, 89)])

        days = data.loc[(data['period_month'] == month) & (data['period_year'] == year)].drop_duplicates(
            ['work_calendar_daytype', 'work_calendar_day'])
        cls.days_types = dict(zip(days['work_calendar_day'].to_list(), days['work_calendar_daytype'].to_list()))

        # Список с выходными днями (нужен ниже для формул)
        list_weekend = []
        for day, type_day in cls.days_types.items():
            if type_day == 1:
                list_weekend.append(columns[day])

        # делаем датафрейм по одному месяцу
        month_year = data.loc[(data['period_month'] == month) & (data['period_year'] == year)]

        # получаем список работников из верхнего фрейма
        workers = month_year.drop_duplicates(['job_name', 'job_department',
                                              'job_position', 'period_month',
                                              'period_year'])

        # бежим по датафрейму работников
        dataframe_worker = []
        count_1 = 6
        for _, worker in workers.iterrows():
            # тут содержится вся инфа на работника
            worker_to_insert = month_year.loc[(month_year['job_name'] == worker['job_name']) &
                                              (month_year['job_department'] == worker['job_department']) &
                                              (month_year['job_position'] == worker['job_position']) &
                                              (month_year['period_month'] == worker['period_month']) &
                                              (month_year['period_year'] == worker['period_year'])]

            # тут фрейм на одного работника
            worker_to_insert = worker_to_insert.sort_values(by='work_calendar_day')

            # счётчик - нужен для формул ниже
            count_shift = count_1

            # лист, куда будем добавлять подготовленные строки
            rows_worker = []

            # обрабатываем построчно каждую запись фрейма работника
            for _, ins in worker_to_insert.iterrows():
                weekend_flag = True
                row = pd.DataFrame()

                # заполняем массив, который содержит координаты ячеек где нужно поставить морковку
                if ins['event'] == 'correction':
                    cls.correction_list.append((count_shift + 1, ins['work_calendar_day'] + 8))

                # Гиперссылка на карточку сотрудника
                hyperlink_name = (f'=HYPERLINK("{jira_server}/issues/?jql=summary~%27{ins["job_name"]}%27", '
                                  f'"{ins["job_name"]}")')

                # Гиперссылка на карточку отдела
                hyperlink_department = (f'=HYPERLINK("{jira_server}/issues/?jql=status=Трудоустройство and '
                                        f'cf[14829]=%27{ins["job_department"]}%27", "{ins["job_department"]}")')

                # Ставим гиперссылку имени на строку с базовым контрактом
                row['B'] = [hyperlink_name] if ins['kontrakt_name'][0] == 'О' else [ins['job_name']]

                # Ставим гиперссылку отдела на строку с базовым контрактом
                row['C'] = [hyperlink_department] if ins['kontrakt_name'][0] == 'О' else [ins['job_department']]

                row['D'] = [ins['job_position']]
                row['E'] = [ins['kontrakt_name']]
                row['F'] = ['День' if ins['day_night'] == 'day' else 'Ночь']
                row['G'] = ['К-ка' if ins['event'] == 'trip' else ' ']
                row['H'] = ['П-ка' if ins['event'] == 'permit' else ' ']

                # Тут выставляем всё то, у чего есть timetracking в будние дни. И ставим гиперссылку
                if (ins['event'] in ['shift', 'trip', 'permit', 'correction'] and ins['work_time'] > 0
                        and ins['work_calendar_daytype'] == 0 and ins['kontrakt_filter'] != '0'
                        and ins['kontrakt_timetracking'] > 0):
                    value = (f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", '
                             f'{ins["kontrakt_timetracking"]})')

                # Выставляем shift по заявке в выходной день, если есть filter и worktime > 0.
                # Например: сменщик работает в воскресенье и выполнял заявки
                elif (ins['event'] == 'shift' and ins['work_time'] > 0 and ins['work_calendar_daytype'] == 1
                      and ins['kontrakt_filter'] != '0' and ins['kontrakt_timetracking'] > 0):
                    value = (f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", '
                             f'{ins["kontrakt_timetracking"]})')

                # Выставляем командировку на выходных. Если у человека командировка захватывает выходные,
                # то это считается выходным, но человек всё еще в командировке, так что нам необходима тут гиперссылка
                elif ins['event'] == 'trip' and ins['work_calendar_daytype'] == 1:
                    value = f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", "В")'
                    weekend_flag = False

                # Выставляем permit в выходной день
                elif ins['event'] == 'permit' and ins['work_calendar_daytype'] == 1:
                    value = (f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", '
                             f'{ins["kontrakt_timetracking"]})')

                # Если событие - корректировка и work_time == 0, то значит человек не явился на работу, ставим Н
                elif ins['work_time'] == 0 and ins['event'] == 'correction' and ins["kontrakt_name"][0] == 'О':
                    value = 'Н'

                # Ставим отпуск с гиперссылкой на DOCCORP
                elif ins['event'] == 'otpusk':
                    value = f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", "О")'

                # Ставим отгул с гиперссылкой на DOCCORP
                elif ins['event'] == 'compensatory':
                    value = f'=HYPERLINK("{jira_server}/issues/?jql=issue in ({ins["kontrakt_filter"]})", "От")'

                # Ставим Б, если событие больничный
                elif ins['event'] == 'hospital':
                    value = 'Б'

                # Обрабатываем выходные дни
                elif (ins['work_calendar_daytype'] == 1 and ins['kontrakt_name'][0] == 'О' and ins['work_time'] == 0
                      and weekend_flag):
                    value = 'В'

                # Выставляем всё остальное (как правило - это время списанное на базовый контракт)
                else:
                    value = ins['kontrakt_timetracking'] if ins['kontrakt_timetracking'] > 0 else ' '

                    
                # присваиваем значение
                row[f'{str(ins["work_calendar_day"])}'] = value

                # добавляем в список, для дальнейшей конкатенации
                rows_worker.append(row)

            # конкатенируем строки работника в один фрейм и группируем
            df_rows = (pd.concat(rows_worker, ignore_index=True).
                       groupby(['E', 'F', 'G', 'H'], as_index=False).last().sort_values('E', ascending=False))

            # добиваем пустыми данными дни
            for column in columns.keys():
                if str(column) not in df_rows.columns.values.tolist():
                    df_rows[f'{str(column)}'] = ' '

            # Создаем колонки с AN по AX (Там будут формулы)
            for column in [f'A{chr(i)}' for i in range(78, 89)]:
                df_rows[f'{column}'] = ' '

            # счетчик для формул справа
            count_right = count_1 + 1

            # ставим формулы
            for index, ins in df_rows.iterrows():
                # Количество смен (день)
                df_rows.loc[index, 'AN'] = f'=COUNT(I{count_right}:AM{count_right})' if ins["F"] == 'День' else ' '

                # Количество смен (ночь)
                df_rows.loc[index, 'AO'] = f'=COUNT(I{count_right}:AM{count_right})' if ins["F"] == 'Ночь' else ' '

                # Отпуск
                df_rows.loc[index, 'AP'] = f'=COUNTIF(I{count_right}:AM{count_right}, "О")' if ins["E"][
                                                                                                   0] == 'О' else ' '
                # Отгул
                df_rows.loc[index, 'AQ'] = f'=COUNTIF(I{count_right}:AM{count_right}, "От")' if ins['E'][
                                                                                                    0] == 'О' else ' '
                # Больничный
                df_rows.loc[index, 'AR'] = f'=COUNTIF(I{count_right}:AM{count_right}, "Б")' if ins['E'][
                                                                                                   0] == 'О' else ' '
                # Неявка
                df_rows.loc[index, 'AS'] = f'=COUNTIF(I{count_right}:AM{count_right}, "Н")' if ins['E'][
                                                                                                   0] == 'О' else ' '
                # Кол-во командировок (Если человек на выходных в командировке - то же считаем)
                df_rows.loc[index, 'AT'] = (f'=COUNTIF(I{count_right}:AM{count_right}, "В") + '
                                            f'SUM(I{count_right}:AM{count_right})/8') if ins['G'] == 'К-ка' else ' '

                # Кол-во отработанных часов (день)
                df_rows.loc[index, 'AU'] = f'=SUM(I{count_right}:AM{count_right})' if ins['F'] == 'День' else ' '

                # Кол-во отработанных часов (ночь)
                df_rows.loc[index, 'AV'] = f'=SUM(I{count_right}:AM{count_right})' if ins['F'] == 'Ночь' else ' '

                # собираем формулу для выходного дня
                formula_weekend = f'=SUM('
                temp_formula = ''
                for i in list_weekend:
                    temp_formula = temp_formula + f'{i}{count_right},'
                formula_weekend = formula_weekend + temp_formula + ')'
                formula_weekend = ' ' if formula_weekend == '=SUM()' else formula_weekend

                # Ставим формулу
                df_rows.loc[index, 'AW'] = formula_weekend if ins['F'] == 'День' else ' '
                df_rows.loc[index, 'AX'] = formula_weekend if ins['F'] == 'Ночь' else ' '

                count_right += 1
            ###########################################################################
            # тут начинаем работать со строкой total

            total_row = pd.DataFrame()
            total_row['B'] = [worker['job_name']]
            total_row['C'] = [worker['job_department']]
            total_row['D'] = [worker['job_position']]
            total_row['E'] = ['ИТОГО:']
            total_row['F'] = [' ']
            total_row['G'] = [' ']
            total_row['H'] = [' ']

            # Возимся с координатами, нужно ниже для формул
            count_1 = len(df_rows) + 1 + count_1

            # получаем дни работника в список
            days = worker_to_insert['work_calendar_day'].drop_duplicates().to_list()

            # Ставим формул в строку ИТОГО. (Сумма по столбцу)
            for day in days:
                total_row[f'{str(day)}'] = [f'=SUM({columns[day]}{count_1 - 1}:{columns[day]}{count_1 - len(df_rows)})']

                # Эту конструкцию надо оставить (Вдруг не надо будет считать сумму по столбцу)
                # total_row[f'{str(day)}'] = float(
                #     worker_to_insert['work_time'].loc[worker_to_insert['work_calendar_day'] == day].iloc[0])

            # добиваем колонки пустыми данными
            for column in columns.keys():
                if str(column) not in total_row.columns.values.tolist():
                    total_row[f'{str(column)}'] = ' '

            # создаем колонки с AN по AX (Тут будут формулы строки ИТОГО)
            for column in [f'A{chr(i)}' for i in range(78, 89)]:
                total_row[f'{column}'] = ' '

            # Собираем все ячейки для формул Отработано смен (день/ночь) в строку ИТОГО
            formula_working_shift_day = defaultdict(list)
            formula_working_shift_night = defaultdict(list)
            for _, row in df_rows.iterrows():
                for key, val in columns.items():
                    if row[str(key)] not in ['О', 'Б', 'В', 'Н', 'От', ' ', ''] and row['F'] == 'День':
                        formula_working_shift_day[key].append(f'{val}{count_shift + 1}')
                    if row[str(key)] not in ['О', 'Б', 'В', 'Н', 'От', ' ', ''] and row['F'] == 'Ночь':
                        formula_working_shift_night[key].append(f'{val}{count_shift + 1}')
                count_shift += 1

            # Тут собираем формулу (день)
            res = ''
            for key, val in formula_working_shift_day.items():
                temp_form = ''
                for i in val:
                    temp_form = temp_form + str(i) + ','
                res = res + 'IF(SUM(' + temp_form + ')>0,1,"FALSE"),'
            formula_shift_day = '=COUNT(' + res[:-1] + ')'
            formula_shift_day = '' if formula_shift_day == '=COUNT()' else formula_shift_day

            # Тут собираем формулу (ночь)
            res = ''
            for key, val in formula_working_shift_night.items():
                temp_form = ''
                for i in val:
                    temp_form = temp_form + str(i) + ','
                res = res + 'IF(SUM(' + temp_form + ')>0,1,"FALSE"),'
            formula_shift_night = '=COUNT(' + res[:-1] + ')'
            formula_shift_night = '' if formula_shift_night == '=COUNT()' else formula_shift_night

            # Расставляем формулы по ячейкам
            total_row['AN'] = [formula_shift_day]
            total_row['AO'] = [formula_shift_night]
            total_row['AP'] = [f'=MAX(AP{count_1 - 1}:AP{count_1 - len(df_rows)})']
            total_row['AQ'] = [f'=MAX(AQ{count_1 - 1}:AQ{count_1 - len(df_rows)})']
            total_row['AR'] = [f'=MAX(AR{count_1 - 1}:AR{count_1 - len(df_rows)})']
            total_row['AS'] = [f'=MAX(AS{count_1 - 1}:AS{count_1 - len(df_rows)})']

            total_row['AT'] = [f'=SUM(AT{count_1 - 1}:AT{count_1 - len(df_rows)})']
            total_row['AU'] = [f'=SUM(AU{count_1 - 1}:AU{count_1 - len(df_rows)})']
            total_row['AV'] = [f'=SUM(AV{count_1 - 1}:AV{count_1 - len(df_rows)})']
            total_row['AW'] = [f'=SUM(AW{count_1 - 1}:AW{count_1 - len(df_rows)})']
            total_row['AX'] = [f'=SUM(AX{count_1 - 1}:AX{count_1 - len(df_rows)})']

            # Объединяем строки и ИТОГО
            df_with_total = pd.concat([total_row, df_rows], ignore_index=True).sort_values(by='E', ascending=False)
            # добавляем в список с другими работниками
            dataframe_worker.append(df_with_total.sort_values(by='E', ascending=False))

        # Конкатенируем в общий фрейм
        one_month = pd.concat(dataframe_worker, ignore_index=True)
        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=Finished form a dataframe {month} {year}')

        # Делаем правильный порядок колонок
        one_month = one_month[columns_dop]

        # сбрасываем индексы
        one_month = one_month.reset_index()

        return one_month

    @classmethod
    def doc_header(cls, wb, month, year):
        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=Started creating sheet {month} {year}')

        ws = wb.create_sheet(month)

        ws.freeze_panes = 'A7'
        # устанавливаем фильтры для полей
        ws.auto_filter.ref = 'A6:H6'
        ws['B3'] = "ТАБЕЛЬ учета использования рабочего времени"
        ws['D3'] = month + ' ' + str(year)

        ws['B3'].font = openpyxl.styles.Font(
            name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
        ws['D3'].font = openpyxl.styles.Font(
            name='Arial Cyr', charset=204, family=2.0, b=True, sz=18.0)

        ws.merge_cells('J1:AE1')
        ws.merge_cells('J2:AE2')
        ws.merge_cells('J3:AE3')
        ws.merge_cells('J4:AE4')
        medium_border = Border(left=Side(style='medium'),
                               right=Side(style='medium'),
                               top=Side(style='medium'),
                               bottom=Side(style='medium'))
        ws.cell(row=4, column=10).border = medium_border

        for r in range(1, 5):
            ws.cell(row=r, column=9).border = Border(left=Side(style='medium'))
            ws.cell(row=r, column=31).border = Border(
                right=Side(style='medium'))
        for c in range(9, 32):
            ws.cell(row=4, column=c).border = Border(
                bottom=Side(style='medium'))
        ws.cell(row=4, column=9).border = Border(
            left=Side(style='medium'), bottom=Side(style='medium'))
        ws.cell(row=4, column=31).border = Border(
            right=Side(style='medium'), bottom=Side(style='medium'))

        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        yellow_font = openpyxl.styles.Font(
            name='Arial Cyr', charset=204, family=2.0, b=True, sz=16.0)

        ws['I1'] = "О"
        ws['I2'] = "Б"
        ws['I3'] = "В"
        ws['I4'] = "Н"
        ws['I1'].font = yellow_font
        ws['I1'].fill = yellow_fill

        ws['I2'].font = yellow_font
        ws['I2'].fill = yellow_fill

        ws['I3'].font = yellow_font
        ws['I3'].fill = yellow_fill

        ws['I4'].font = yellow_font
        ws['I4'].fill = yellow_fill

        ws['J1'] = "Отпуск (заполняется только в белой строке)"
        ws['J2'] = "Больничный (заполняется только в белой строке)"
        ws['J3'] = "Выходной (заполняется только в белой строке)"
        ws['J4'] = "Неявка по невыясненным причинам  (заполняется только в белой строке)"

        ws['J1'].fill = yellow_fill
        ws['J2'].fill = yellow_fill
        ws['J3'].fill = yellow_fill
        ws['J4'].fill = yellow_fill

        # ================ ставим морковку ============================================
        ws.merge_cells('AG1:AP4')

        for r in range(1, 5):
            ws.cell(row=r, column=33).border = Border(left=Side(style='medium'))
            ws.cell(row=r, column=42).border = Border(
                right=Side(style='medium'))

        for c in range(33, 43):
            ws.cell(row=4, column=c).border = Border(
                bottom=Side(style='medium'))

        ws.cell(row=4, column=33).border = Border(
            left=Side(style='medium'), bottom=Side(style='medium'))
        ws.cell(row=4, column=42).border = Border(
            right=Side(style='medium'), bottom=Side(style='medium'))

        ws['AG1'] = '"Морковным" цветом выделены дни, в которых была корректировка рабочего времени'
        ws['AG1'].fill = PatternFill(
            start_color="F79646", end_color="F79646", fill_type="solid")

        ws['AG1'].font = openpyxl.styles.Font(
            name='Arial Cyr', charset=204, family=2.0, b=False, sz=12.0)
        ws['AG1'].alignment = Alignment(
            wrap_text=True, vertical='center', horizontal='center')

        # =============================================================================

        ws['A6'] = "№"
        ws['B6'] = "Ф.И.О"
        ws['C6'] = "Подразделение"
        ws['D6'] = "специальность, профессия"
        ws['E6'] = "Контракт"
        ws['F6'] = "Н/Д"
        ws['G6'] = "Командировка"
        ws['H6'] = "Переработка"

        column_font = openpyxl.styles.Font(
            name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
        column_alignment = Alignment(
            wrap_text=True, vertical='center', horizontal='center')
        column_alignment_rotation = Alignment(
            textRotation=90, wrap_text=True, vertical='center', horizontal='center')

        ws['A6'].font = column_font
        ws['A6'].alignment = column_alignment
        ws.column_dimensions['A'].width = value = 4.57

        ws['B6'].font = column_font
        ws['B6'].alignment = column_alignment
        ws.column_dimensions['B'].width = value = 32.57

        ws['C6'].font = column_font
        ws['C6'].alignment = column_alignment
        ws.column_dimensions['C'].width = value = 32.57

        ws['D6'].font = column_font
        ws['D6'].alignment = column_alignment
        ws.column_dimensions['D'].width = value = 22.14

        ws['E6'].font = column_font
        ws['E6'].alignment = column_alignment
        ws.column_dimensions['E'].width = value = 22.14

        for col in range(6, 9):
            ws.cell(row=6, column=col).font = column_font
            ws.cell(row=6, column=col).alignment = column_alignment_rotation
            ws.column_dimensions[get_column_letter(col)].width = value = 5

        for col in range(9, 40):
            ws.cell(row=6, column=col).font = column_font
            ws.column_dimensions[get_column_letter(col)].width = value = 3.7

        ws['AN6'] = "Отработано смен (день)"
        ws['AO6'] = "Отработано смен (ночь)"
        ws['AP6'] = "Отпуск"
        ws['AQ6'] = "Отгул"
        ws['AR6'] = "Больничный"
        ws['AS6'] = "Неявка"
        ws['AT6'] = "Командировка"
        ws['AU6'] = "Отработано часы (день)"
        ws['AV6'] = "Отработано часы (ночь)"
        ws['AW6'] = "Отработано в выходные часы (день)"
        ws['AX6'] = "Отработано в выходные часы (ночь)"

        column_font = openpyxl.styles.Font(name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
        right_column_alignment = Alignment(textRotation=90)
        beige_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        for col in range(40, 51):
            ws.cell(row=6, column=col).font = column_font
            ws.cell(row=6, column=col).alignment = right_column_alignment
            ws.column_dimensions[get_column_letter(col)].width = value = 5
            ws.cell(row=6, column=col).fill = beige_fill if col > 45 else green_fill

    @classmethod
    def beauty(cls, wb, one_month, month, year):
        ws = wb[month]

        # выгружаем pandas dataframe на лист
        for row in dataframe_to_rows(one_month, index=False, header=False):
            ws.append(row)

        total_fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
        font_bold = Font(bold=True)

        # раскрашиваем строку total
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == 'ИТОГО:':
                    for c in row:
                        c.font = font_bold
                        c.fill = total_fill

        # раскрашиваем выходные дни
        for day, type_day in cls.days_types.items():
            ws.cell(row=6, column=8 + day).value = day
            if type_day == 1:
                ws.cell(row=6, column=8 + day).fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid")

        cls.days_types = {}

        # делаем сетку на всём документе
        thin_border1 = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))
        last_empty_row = len(list(ws.rows)) + 1
        for row in range(6, last_empty_row):
            for col in range(1, 51):
                if col == 1 and row > 6:
                    ws.cell(row=row, column=1, value=row - 6)
                ws.cell(row=row, column=col).border = thin_border1
                if col > 39 and row > 6:
                    ws.cell(row=row, column=col).font = font_bold

        # раскрашиваем гиперссылки
        hyper = Font(color='0000ff')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if str(cell.value) and str(cell.value)[:6] == '=HYPER':
                    cell.font = hyper

        # красим корректировки
        # print(cls.correction_list)
        for cor in cls.correction_list:
            ws.cell(row=cor[0], column=cor[1]).fill = PatternFill(
                start_color="F79646", end_color="F79646", fill_type="solid")

        requests.post(
            f'{settings.SERVICE_REST}/service/log?level={logging.INFO}&message=Finished creating sheet {month} {year}')
